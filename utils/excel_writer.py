"""
Generador de Excel de evidencia de revisión tributaria.
Exporta el registro con validaciones marcadas, coloreadas y con firma de auditor.
"""

import io
from datetime import datetime
import pandas as pd
import openpyxl
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows

# Colores
COLOR_OK = "C6EFCE"          # Verde claro
COLOR_ALERTA = "FFEB9C"      # Amarillo
COLOR_ERROR = "FFC7CE"       # Rojo claro
COLOR_HEADER = "1F4E79"      # Azul oscuro
COLOR_HEADER_VALIDACION = "4472C4"  # Azul medio
COLOR_SUBHEADER = "2F75B6"
COLOR_BLANCO = "FFFFFF"
COLOR_GRIS = "F2F2F2"

FONT_HEADER = Font(name="Calibri", bold=True, color="FFFFFF", size=10)
FONT_NORMAL = Font(name="Calibri", size=9)
FONT_BOLD = Font(name="Calibri", bold=True, size=9)
FONT_TITULO = Font(name="Calibri", bold=True, size=12, color="1F4E79")

BORDER_THIN = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)

FILL_OK = PatternFill(start_color=COLOR_OK, end_color=COLOR_OK, fill_type="solid")
FILL_ALERTA = PatternFill(start_color=COLOR_ALERTA, end_color=COLOR_ALERTA, fill_type="solid")
FILL_ERROR = PatternFill(start_color=COLOR_ERROR, end_color=COLOR_ERROR, fill_type="solid")
FILL_HEADER = PatternFill(start_color=COLOR_HEADER, end_color=COLOR_HEADER, fill_type="solid")
FILL_HEADER_VAL = PatternFill(start_color=COLOR_HEADER_VALIDACION, end_color=COLOR_HEADER_VALIDACION, fill_type="solid")
FILL_GRIS = PatternFill(start_color=COLOR_GRIS, end_color=COLOR_GRIS, fill_type="solid")


def _apply_fill_by_estado(cell, estado: str):
    if estado == "OK":
        cell.fill = FILL_OK
    elif estado == "ALERTA":
        cell.fill = FILL_ALERTA
    elif estado == "ERROR":
        cell.fill = FILL_ERROR


def generar_excel_evidencia_compras(
    df_original: pd.DataFrame,
    resultados_validacion: list,
    empresa_nombre: str,
    empresa_ruc: str,
    periodo: str,
    contador_nombre: str,
    supervisor_nombre: str = "",
    fecha_revision: datetime = None,
) -> bytes:
    """
    Genera un archivo Excel de evidencia con:
    - Portada con datos de la revisión
    - Hoja de Registro de Compras con validaciones coloreadas
    - Hoja de Resumen de Observaciones
    - Hoja de Estadísticas

    Retorna bytes del archivo Excel.
    """
    if fecha_revision is None:
        fecha_revision = datetime.now()

    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # Remover hoja default

    # ─── HOJA 1: PORTADA ────────────────────────────────────
    ws_portada = wb.create_sheet("PORTADA")
    ws_portada.column_dimensions["A"].width = 30
    ws_portada.column_dimensions["B"].width = 50

    datos_portada = [
        ("EVIDENCIA DE REVISIÓN TRIBUTARIA", None),
        ("", None),
        ("EMPRESA:", empresa_nombre),
        ("RUC:", empresa_ruc),
        ("PERÍODO:", periodo),
        ("MÓDULO:", "Registro de Compras - IGV"),
        ("", None),
        ("CONTADOR REVISOR:", contador_nombre),
        ("SUPERVISOR:", supervisor_nombre or "Pendiente de aprobación"),
        ("FECHA DE REVISIÓN:", fecha_revision.strftime("%d/%m/%Y %H:%M")),
        ("", None),
        ("GENERADO POR:", "Sistema de Revisión Tributaria"),
        ("VERSIÓN:", "1.0"),
    ]

    for i, (etiqueta, valor) in enumerate(datos_portada, start=1):
        celda_etiq = ws_portada.cell(row=i, column=1, value=etiqueta)
        if i == 1:
            celda_etiq.font = FONT_TITULO
            ws_portada.merge_cells(f"A{i}:B{i}")
            celda_etiq.alignment = Alignment(horizontal="center")
        else:
            celda_etiq.font = FONT_BOLD
            if valor is not None:
                celda_val = ws_portada.cell(row=i, column=2, value=valor)
                celda_val.font = FONT_NORMAL

    # Firma
    ws_portada.cell(row=16, column=1, value="FIRMA CONTADOR:").font = FONT_BOLD
    ws_portada.cell(row=16, column=2, value="_" * 40)
    ws_portada.cell(row=18, column=1, value="FIRMA SUPERVISOR:").font = FONT_BOLD
    ws_portada.cell(row=18, column=2, value="_" * 40)

    # ─── HOJA 2: REGISTRO DE COMPRAS CON VALIDACIONES ────────
    ws_rc = wb.create_sheet("REG. COMPRAS VALIDADO")

    # Columnas a mostrar del original (las más importantes)
    cols_mostrar = [
        "NUM CORRELATIVO", "FECHA DE EMISIÓN", "TIPO DOC",
        "SERIE", "NRO COMPROBANTE", "NRO DOC PROVEEDOR", "RAZÓN SOCIAL",
        "DESCRIPCIÓN",
        "ADQ GRA DES OPE GRA - BASE", "ADQ GRA DES OPE GRA - IGV",
        "VALOR NO GRABADO", "IMPORTE TOTAL MN", "MONEDA", "TIPO DE CAMBIO",
        "NRO CONSTANCIA DETRACCIÓN",
    ]
    cols_mostrar = [c for c in cols_mostrar if c in df_original.columns]

    # Columnas de validación que agregamos
    cols_validacion = [
        "SEMÁFORO", "ESTADO", "VAL. IGV", "VAL. TIPO CP",
        "VAL. BANCARIZACIÓN", "VAL. DETRACCIÓN", "VAL. SUNAT FE",
        "VAL. RUC", "VAL. TC", "OBSERVACIONES",
    ]

    todas_cols = cols_mostrar + cols_validacion

    # Título
    ws_rc.merge_cells("A1:T1")
    titulo = ws_rc.cell(row=1, column=1,
                        value=f"REGISTRO DE COMPRAS REVISADO - {empresa_nombre} | RUC: {empresa_ruc} | Período: {periodo}")
    titulo.font = FONT_TITULO
    titulo.alignment = Alignment(horizontal="center")
    titulo.fill = FILL_GRIS

    # Subtítulo
    ws_rc.merge_cells("A2:T2")
    subtitulo = ws_rc.cell(row=2, column=1,
                           value=f"Revisión: {contador_nombre} | Fecha: {fecha_revision.strftime('%d/%m/%Y %H:%M')}")
    subtitulo.font = Font(name="Calibri", italic=True, size=9)
    subtitulo.alignment = Alignment(horizontal="center")

    fila_header = 4

    # Headers de datos originales
    for col_idx, col_name in enumerate(todas_cols, start=1):
        cell = ws_rc.cell(row=fila_header, column=col_idx, value=col_name)
        if col_name in cols_validacion:
            cell.fill = FILL_HEADER_VAL
        else:
            cell.fill = FILL_HEADER
        cell.font = FONT_HEADER
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER_THIN

    ws_rc.row_dimensions[fila_header].height = 35

    # Datos
    for row_idx, (_, fila_df) in enumerate(df_original.iterrows(), start=fila_header + 1):
        resultado = resultados_validacion[row_idx - fila_header - 1] if (row_idx - fila_header - 1) < len(resultados_validacion) else {}

        estado = resultado.get("estado", "")
        semaforo = resultado.get("semaforo", "")
        detalle = resultado.get("detalle", {})
        observaciones = "; ".join(resultado.get("observaciones", []))

        # Preparar datos de validación por columna
        def _val_ok(clave):
            r = detalle.get(clave, {})
            if not r:
                return "N/A"
            if r.get("ok") is True:
                return "✔ OK"
            elif r.get("ok") is False:
                return "✘ ERROR"
            return "⚠ REVISAR"

        validaciones_fila = [
            semaforo, estado,
            _val_ok("igv"),
            _val_ok("tipo_cp"),
            _val_ok("bancarizacion"),
            _val_ok("detraccion"),
            resultado.get("sunat_fe", {}).get("estado", "N/A"),
            resultado.get("ruc", {}).get("condicion", "N/A"),
            _val_ok("tc"),
            observaciones,
        ]

        for col_idx, col_name in enumerate(cols_mostrar, start=1):
            val = fila_df.get(col_name, "")
            if val is None:
                val = ""
            cell = ws_rc.cell(row=row_idx, column=col_idx, value=str(val) if not isinstance(val, (int, float)) else val)
            cell.font = FONT_NORMAL
            cell.border = BORDER_THIN
            _apply_fill_by_estado(cell, estado)
            cell.alignment = Alignment(wrap_text=False)

        for col_idx, val in enumerate(validaciones_fila, start=len(cols_mostrar) + 1):
            cell = ws_rc.cell(row=row_idx, column=col_idx, value=val)
            cell.font = FONT_NORMAL
            cell.border = BORDER_THIN
            _apply_fill_by_estado(cell, estado)
            if col_idx == len(cols_mostrar) + len(cols_validacion):  # Observaciones
                cell.alignment = Alignment(wrap_text=True)
            else:
                cell.alignment = Alignment(horizontal="center")

    # Autoajustar anchos
    anchos = {
        "NUM CORRELATIVO": 6, "FECHA DE EMISIÓN": 12, "TIPO DOC": 7,
        "SERIE": 8, "NRO COMPROBANTE": 12, "NRO DOC PROVEEDOR": 13,
        "RAZÓN SOCIAL": 30, "DESCRIPCIÓN": 25,
        "ADQ GRA DES OPE GRA - BASE": 14, "ADQ GRA DES OPE GRA - IGV": 12,
        "VALOR NO GRABADO": 12, "IMPORTE TOTAL MN": 14, "MONEDA": 7, "TIPO DE CAMBIO": 10,
        "NRO CONSTANCIA DETRACCIÓN": 15,
        "SEMÁFORO": 8, "ESTADO": 10, "VAL. IGV": 10, "VAL. TIPO CP": 12,
        "VAL. BANCARIZACIÓN": 14, "VAL. DETRACCIÓN": 14, "VAL. SUNAT FE": 12,
        "VAL. RUC": 12, "VAL. TC": 10, "OBSERVACIONES": 50,
    }
    for col_idx, col_name in enumerate(todas_cols, start=1):
        ws_rc.column_dimensions[get_column_letter(col_idx)].width = anchos.get(col_name, 12)

    ws_rc.freeze_panes = f"A{fila_header + 1}"

    # ─── HOJA 3: RESUMEN DE OBSERVACIONES ────────────────────
    ws_resumen = wb.create_sheet("OBSERVACIONES")

    ws_resumen.merge_cells("A1:E1")
    t = ws_resumen.cell(row=1, column=1, value="RESUMEN DE OBSERVACIONES")
    t.font = FONT_TITULO
    t.alignment = Alignment(horizontal="center")

    headers_res = ["N°", "RAZÓN SOCIAL", "RUC", "COMPROBANTE", "OBSERVACIÓN"]
    for ci, h in enumerate(headers_res, start=1):
        c = ws_resumen.cell(row=3, column=ci, value=h)
        c.fill = FILL_HEADER
        c.font = FONT_HEADER
        c.alignment = Alignment(horizontal="center")
        c.border = BORDER_THIN

    anchos_res = {"A": 5, "B": 35, "C": 13, "D": 20, "E": 70}
    for col_letra, ancho in anchos_res.items():
        ws_resumen.column_dimensions[col_letra].width = ancho

    fila_res = 4
    obs_num = 1
    for row_idx, (_, fila_df) in enumerate(df_original.iterrows()):
        resultado = resultados_validacion[row_idx] if row_idx < len(resultados_validacion) else {}
        observaciones = resultado.get("observaciones", [])
        if not observaciones:
            continue
        razon_social = str(fila_df.get("RAZÓN SOCIAL", ""))[:40]
        ruc = str(fila_df.get("NRO DOC PROVEEDOR", ""))
        serie = str(fila_df.get("SERIE", ""))
        nro = str(fila_df.get("NRO COMPROBANTE", ""))
        comprobante = f"{serie}-{nro}"

        for obs in observaciones:
            estado = resultado.get("estado", "")
            ws_resumen.cell(row=fila_res, column=1, value=obs_num).font = FONT_NORMAL
            ws_resumen.cell(row=fila_res, column=2, value=razon_social).font = FONT_NORMAL
            ws_resumen.cell(row=fila_res, column=3, value=ruc).font = FONT_NORMAL
            ws_resumen.cell(row=fila_res, column=4, value=comprobante).font = FONT_NORMAL
            obs_cell = ws_resumen.cell(row=fila_res, column=5, value=obs)
            obs_cell.font = FONT_NORMAL
            obs_cell.alignment = Alignment(wrap_text=True)

            for ci in range(1, 6):
                c = ws_resumen.cell(row=fila_res, column=ci)
                c.border = BORDER_THIN
                _apply_fill_by_estado(c, estado)

            ws_resumen.row_dimensions[fila_res].height = 25
            fila_res += 1
            obs_num += 1

    if obs_num == 1:
        ws_resumen.cell(row=4, column=1, value="✔ Sin observaciones. Revisión conforme.")

    # ─── HOJA 4: ESTADÍSTICAS ────────────────────────────────
    ws_stats = wb.create_sheet("ESTADÍSTICAS")

    total = len(resultados_validacion)
    errores = sum(1 for r in resultados_validacion if r.get("estado") == "ERROR")
    alertas = sum(1 for r in resultados_validacion if r.get("estado") == "ALERTA")
    ok = sum(1 for r in resultados_validacion if r.get("estado") == "OK")

    ws_stats.merge_cells("A1:C1")
    t = ws_stats.cell(row=1, column=1, value="ESTADÍSTICAS DE REVISIÓN")
    t.font = FONT_TITULO
    t.alignment = Alignment(horizontal="center")

    datos_stats = [
        ("Total comprobantes revisados", total),
        ("✔ Sin observaciones", ok),
        ("⚠ Con alertas", alertas),
        ("✘ Con errores", errores),
        ("% Conformidad", f"{(ok/total*100):.1f}%" if total > 0 else "N/A"),
    ]

    for i, (desc, val) in enumerate(datos_stats, start=3):
        c1 = ws_stats.cell(row=i, column=1, value=desc)
        c1.font = FONT_BOLD
        c1.border = BORDER_THIN
        c2 = ws_stats.cell(row=i, column=2, value=val)
        c2.font = FONT_NORMAL
        c2.border = BORDER_THIN
        c2.alignment = Alignment(horizontal="center")
        if "error" in desc.lower():
            c1.fill = FILL_ERROR
            c2.fill = FILL_ERROR
        elif "alerta" in desc.lower():
            c1.fill = FILL_ALERTA
            c2.fill = FILL_ALERTA
        elif "sin observaciones" in desc.lower():
            c1.fill = FILL_OK
            c2.fill = FILL_OK

    ws_stats.column_dimensions["A"].width = 35
    ws_stats.column_dimensions["B"].width = 20

    # Guardar en bytes
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()
